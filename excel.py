import openpyxl
import os

os.chdir("C:\\Users\\user\\Downloads\\automate\\python_automation")

workbook = openpyxl.load_workbook('example.xlsx')

print(type(workbook))

sheet = workbook['Sheet1']
print(type(sheet))

cell = sheet['A1']

print(cell.value)
print(type(cell.value))

# sheet.cell(row=1, column=2) == sheet['B1']
print(sheet.cell(row=1, column=2))

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)

wb = openpyxl.Workbook()
print(wb.sheetnames)

sheet = wb['Sheet']
print(sheet)

sheet['A1'] = 42
sheet['A2'] = 'Hello'

sheet2 = wb.create_sheet()
print(wb.sheetnames)

sheet2.title = 'My new Sheet Name'
print(wb.sheetnames)

wb.save('example2.xlsx')

wb.create_sheet(index=0, title='My Other Sheet')

wb.save('example2.xlsx')